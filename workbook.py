from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import Store

TEAL = "16A34A"
TEAL_DARK = "15803D"
ZEBRA = "F3FAF5"
BORDER = "D9E2E6"

THIN = Border(bottom=Side(style="thin", color=BORDER))
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=TEAL)
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color=TEAL_DARK)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=TEAL_DARK)
MUTED_FONT = Font(name="Calibri", italic=True, size=10, color="7A8699")
BODY = Font(name="Calibri", size=11, color="1F2937")
VALUE_FONT = Font(name="Calibri", bold=True, size=11, color=TEAL_DARK)


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def _fill_sheet(ws, headers, rows):
    ws.append(headers)
    zebra_fill = PatternFill("solid", fgColor=ZEBRA)
    for r_i, row in enumerate(rows):
        ws.append(row)
        if r_i % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r_i + 2, column=col).fill = zebra_fill
    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, 2),
                            max_col=len(headers)):
        for cell in row:
            cell.font = BODY
            cell.border = THIN
    _style_header(ws, len(headers))
    widths = [len(h) for h in headers]
    for row in rows:
        for i, val in enumerate(row[:len(widths)]):
            widths[i] = max(widths[i], len(str(val)))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 3, 9), 42)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    ws.sheet_view.showGridLines = False


def _section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = SECTION_FONT
    return row + 1


def _mini_table(ws, row, headers, data_rows):
    small_hdr = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = small_hdr
        cell.fill = HEADER_FILL
    row += 1
    for vals in data_rows:
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = BODY
            cell.border = THIN
        row += 1
    return row + 1


def _dashboard(store: Store, wb: Workbook):
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_properties.tabColor = TEAL
    dash.sheet_view.showGridLines = False

    dash.merge_cells("A1:H1")
    dash["A1"].value = "\U0001F43E Vet Clinic Database"
    dash["A1"].font = TITLE_FONT
    dash.row_dimensions[1].height = 30
    dash.merge_cells("A2:H2")
    dash["A2"].value = f"Generated {datetime.now():%Y-%m-%d %H:%M} by Vet XLS Studio"
    dash["A2"].font = MUTED_FONT

    vax_due = store.vaccinations_due(30)
    row = _section(dash, 4, "Overview")
    overview = [
        ("Owners", len(store.owners)),
        ("Pets", len(store.pets)),
        ("Appointments", len(store.appointments)),
        ("Treatments & vaccinations", len(store.treatments)),
        ("Vaccinations due within 30 days", len(vax_due)),
        ("Upcoming appointments", len(store.upcoming_appointments(limit=99999))),
        ("Pets hospitalized", store.hospitalized_count()),
        ("Hospital positions free", store.hospital_free()),
    ]
    for label, value in overview:
        dash.cell(row=row, column=1, value=label).font = BODY
        vc = dash.cell(row=row, column=4, value=value)
        vc.font = VALUE_FONT
        vc.alignment = Alignment(horizontal="right")
        row += 1

    row += 1
    guard = (f"{store.night_vet_name} \u00B7 {store.night_vet_phone}"
             if store.night_vet_name else "-")
    dash.cell(row=row, column=1, value="Night guard").font = BODY
    gc = dash.cell(row=row, column=4, value=guard)
    gc.font = VALUE_FONT
    row += 2

    row = _section(dash, row, "Species mix")
    for species, n in store.species_breakdown().items():
        dash.cell(row=row, column=1, value=species).font = BODY
        vc = dash.cell(row=row, column=4, value=n)
        vc.font = VALUE_FONT
        vc.alignment = Alignment(horizontal="right")
        row += 1

    def appt_rows(items):
        out = []
        for a in items:
            pet = store.get_pet(a.pet_id)
            out.append([a.date, a.time, pet.name if pet else "",
                        store.owner_name(pet), a.vet, a.reason])
        return out

    row += 1
    row = _section(dash, row, "Next appointments")
    row = _mini_table(dash, row,
                      ["Date", "Time", "Pet", "Owner", "Vet", "Reason"],
                      appt_rows(store.upcoming_appointments(limit=10)))

    row = _section(dash, row, "Vaccinations due (next 30 days)")
    vax_rows = []
    for due, t in vax_due[:15]:
        pet = store.get_pet(t.pet_id)
        vax_rows.append([t.next_due, pet.name if pet else "",
                         store.owner_name(pet), t.description, t.vet])
    row = _mini_table(dash, row,
                      ["Due date", "Pet", "Owner", "Vaccine", "Vet"], vax_rows)

    for i, w in enumerate([26, 12, 16, 20, 14, 26, 4, 10], start=1):
        dash.column_dimensions[get_column_letter(i)].width = w


def _entity_sheets(store: Store, wb: Workbook):
    ws = wb.create_sheet("Owners")
    ws.sheet_properties.tabColor = TEAL
    rows = [[o.id, o.name, o.phone, o.email, o.address]
            for o in sorted(store.owners, key=lambda x: x.name.lower())]
    _fill_sheet(ws, ["ID", "Name", "Phone", "Email", "Address"], rows)

    ws = wb.create_sheet("Pets")
    ws.sheet_properties.tabColor = TEAL
    rows = [[p.id, p.name, p.species, p.breed, p.sex, p.birth_date,
             p.microchip, store.owner_name(p)]
            for p in sorted(store.pets, key=lambda x: x.name.lower())]
    _fill_sheet(ws, ["ID", "Name", "Species", "Breed", "Sex", "Birth date",
                     "Microchip", "Owner"], rows)

    ws = wb.create_sheet("Appointments")
    ws.sheet_properties.tabColor = TEAL
    rows = [[a.id, a.date, a.time, store.pet_label(a.pet_id), a.vet,
             a.reason, a.status]
            for a in sorted(store.appointments, key=lambda x: (x.date, x.time))]
    _fill_sheet(ws, ["ID", "Date", "Time", "Pet", "Vet", "Reason", "Status"],
                rows)

    ws = wb.create_sheet("Treatments & Vaccinations")
    ws.sheet_properties.tabColor = TEAL
    rows = [[t.id, t.date, t.type, store.pet_label(t.pet_id), t.description,
             t.vet, t.next_due]
            for t in sorted(store.treatments, key=lambda x: x.date)]
    _fill_sheet(ws, ["ID", "Date", "Type", "Pet", "Description", "Vet",
                     "Next due"], rows)


def generate_workbook(store: Store, path: str) -> str:
    wb = Workbook()
    _dashboard(store, wb)
    _entity_sheets(store, wb)
    wb.save(path)
    return path
